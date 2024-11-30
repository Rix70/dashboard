from flask import Blueprint, render_template, jsonify, request, redirect, url_for, session, send_file, flash
import app
from app.models import Parameters, HourlyValues, PageSettings, User, PageHierarchy
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from io import BytesIO
from sqlalchemy import func, case, and_
import pytz
from flask_login import login_user, logout_user, login_required, current_user
from urllib.parse import urlparse
from app import db

# Создаем Blueprint
main = Blueprint('main', __name__)

def convert_time(dt):
    if dt.tzinfo is None:
        dt = pytz.UTC.localize(dt)
    moscow_tz = pytz.timezone('Europe/Moscow')
    return dt.astimezone(moscow_tz)

def get_shift_for_datetime(target_date):
    """Определяет какая смена работает в заданное время"""
    base_date = datetime(2024, 11, 1)  # Опорная дата
    days_diff = (target_date.date() - base_date.date()).days
    
    # Начальное состояние смен на 01.11.2024
    initial_state = {
        'Г': 0,  # дневная смена
        'А': 1,  # ночная смена
        'В': 3,  # выходной
        'Б': 2   # отсыпной
    }
    
    hour = target_date.hour
    is_day_shift = 8 <= hour < 20
    
    for shift_letter, initial_pos in initial_state.items():
        current_pos = (initial_pos + days_diff) % 4
        if current_pos == 0 and is_day_shift:  # Дневная смена
            return shift_letter
        elif current_pos == 1 and not is_day_shift:  # Ночная смена
            return shift_letter
    return None

@main.route('/', methods=['GET', 'POST'])
@login_required
def index():
    # Получаем все параметры для выпадающего списка
    parameters = Parameters.query.all()
    
    # Получаем выбранный параметр из сессии или используем параметр 1 по умолчанию
    selected_parameter_id = session.get('selected_parameter_id', 1)
    
    # Получаем даты из сессии или используем текущую дату
    end_date = session.get('end_date', datetime.now())
    start_date = session.get('start_date', end_date - timedelta(days=7))
    
    # Если даты пришли как строки, преобразуем их
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S%z')
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S%z')
    
    # Обновляем запрос для использования выбранного параметра
    shift_data = db.session.query(
        func.sum(HourlyValues.Value).label('total_value'),
        func.date(HourlyValues.DateTime).label('date'),
        func.extract('hour', HourlyValues.DateTime).label('hour')
    ).filter(
        and_(
            HourlyValues.DateTime.between(start_date, end_date),
            HourlyValues.CodeId == selected_parameter_id
        )
    ).group_by(
        func.date(HourlyValues.DateTime),
        func.extract('hour', HourlyValues.DateTime)
    ).all()
    
    # Агрегируем данные по сменам
    shift_totals = {'А': 0, 'Б': 0, 'В': 0, 'Г': 0}
    
    for value, date, hour in shift_data:
        # Преобразуем date в объект datetime.date, если это строка
        if isinstance(date, str):
            date = datetime.strptime(date, '%Y-%m-%d').date()
        
        # Создаем объект datetime
        dt = datetime.combine(date, datetime.min.time().replace(hour=int(hour)))
        shift = get_shift_for_datetime(dt)
        if shift and value:
            shift_totals[shift] += value
    
    pages = PageSettings.query.all()
    return render_template('index.html', 
                         shift_totals=shift_totals, 
                         pages=pages,
                         start_date=start_date,
                         end_date=end_date,
                         parameters=parameters,
                         selected_parameter_id=selected_parameter_id)

@main.route('/update_dates', methods=['POST'])
def update_dates():
    if request.method == 'POST':
        try:
            start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%dT%H:%M')
            end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%dT%H:%M')
            
            # Конвертируем в московское время
            start_date = convert_time(start_date)
            end_date = convert_time(end_date)
            
            # Сохраняем даты в сессию
            session['start_date'] = start_date
            session['end_date'] = end_date
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': True})
                
            return redirect(request.referrer or url_for('main.index'))
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'error': str(e)}), 400
            flash('Ошибка при обновлении дат', 'error')
            return redirect(request.referrer or url_for('main.index'))

@main.route('/page/<int:page_id>')
def dynamic_page(page_id):
    page_settings = PageSettings.query.get_or_404(page_id)
    start_date = session.get('start_date')
    end_date = session.get('end_date')
    
    # Получаем параметры в порядке из id_list
    parameters = []
    for param_id in page_settings.page_ids:  # используем существующий порядок
        param = Parameters.query.get(param_id)
        if param:
            parameters.append(param)
    
    # Получаем почасовые данные
    hourly_data = db.session.query(
        HourlyValues.DateTime,
        HourlyValues.Value,
        Parameters.Name,
        Parameters.CodeId
    ).join(
        Parameters,
        HourlyValues.CodeId == Parameters.CodeId
    ).filter(
        HourlyValues.DateTime.between(start_date, end_date),
        HourlyValues.CodeId.in_(page_settings.page_ids)
    ).order_by(
        HourlyValues.DateTime
    ).all()
    
    # Группируем почасовые данные
    grouped_hourly_data = {}
    for record in hourly_data:
        date_key = record.DateTime
        if date_key not in grouped_hourly_data:
            grouped_hourly_data[date_key] = {}
        grouped_hourly_data[date_key][record.CodeId] = f"{record.Value:,.2f}".replace(',', ' ').replace('.', ',') if record.Value is not None else '-'

    # Получаем агрегированные суточные данные с учетом типа агрегации
    daily_data = []
    
    for param in parameters:
        if param.aggregation_type == 'weighted_avg' and param.weight_parameter_id:
            # Создаем подзапрос для весовых значений
            weight_values = db.session.query(
                HourlyValues.DateTime,
                HourlyValues.Value.label('weight_value')
            ).filter(
                HourlyValues.CodeId == param.weight_parameter_id,
                HourlyValues.DateTime.between(start_date, end_date),
                HourlyValues.Value != 0  # Исключаем нулевые веса
            ).subquery()

            # Запрос для средневзвешенного с проверкой деления на ноль
            weighted_query = db.session.query(
                func.date(HourlyValues.DateTime).label('date'),
                case(
                    (func.sum(weight_values.c.weight_value) == 0, None),
                    else_=(func.sum(HourlyValues.Value * weight_values.c.weight_value) / 
                          func.sum(weight_values.c.weight_value))
                ).label('aggregation'),
                HourlyValues.CodeId
            ).join(
                Parameters,
                HourlyValues.CodeId == Parameters.CodeId
            ).join(
                weight_values,
                HourlyValues.DateTime == weight_values.c.DateTime
            ).filter(
                HourlyValues.DateTime.between(start_date, end_date),
                HourlyValues.CodeId == param.CodeId
            ).group_by(
                func.date(HourlyValues.DateTime),
                HourlyValues.CodeId
            ).all()

            daily_data.extend([{
                'date': record.date,
                'CodeId': record.CodeId,
                'aggregation': record.aggregation
            } for record in weighted_query])
        else:
            # Запрос для других типов агрегации
            regular_query = db.session.query(
                func.date(HourlyValues.DateTime).label('date'),
                case(
                    (Parameters.aggregation_type == 'sum', func.sum(HourlyValues.Value)),
                    (Parameters.aggregation_type == 'min', func.min(HourlyValues.Value)),
                    (Parameters.aggregation_type == 'max', func.max(HourlyValues.Value)),
                    else_=func.avg(HourlyValues.Value)
                ).label('aggregation'),
                HourlyValues.CodeId
            ).join(
                Parameters,
                HourlyValues.CodeId == Parameters.CodeId
            ).filter(
                HourlyValues.DateTime.between(start_date, end_date),
                HourlyValues.CodeId == param.CodeId
            ).group_by(
                func.date(HourlyValues.DateTime),
                HourlyValues.CodeId
            ).all()
            
            daily_data.extend([{
                'date': record.date,
                'CodeId': record.CodeId,
                'aggregation': record.aggregation
            } for record in regular_query])

    # Группируем данные
    grouped_daily_data = {}
    for record in daily_data:
        date_key = record['date']
        if date_key not in grouped_daily_data:
            grouped_daily_data[date_key] = {}
        grouped_daily_data[date_key][record['CodeId']] = {
            'aggregation': f"{record['aggregation']:,.2f}".replace(',', ' ').replace('.', ',') if record['aggregation'] is not None else '-'
        }

    return render_template('page.html',
                         data=grouped_hourly_data,
                         daily_data=grouped_daily_data,
                         parameters=parameters,
                         page_settings=page_settings)

@main.context_processor
def utility_processor():
    def get_pages():
        return PageSettings.query.all()
    return dict(pages=get_pages())

def calculate_shift_schedule(start_date, days):
    shifts = {
        'А': [], 'Б': [], 'В': [], 'Г': []
    }
    
    # Начальное смещение для каждой смены на 01.11.2024
    # 0 - дневная, 1 - ночная, 2 - отсыпной, 3 - выходной
    initial_state = {
        'Г': 0,  # дневная смена
        'А': 1,  # ночная смена
        'В': 3,  # выходной (на следующий день будет дневная)
        'Б': 2   # отсыпной (после будет выходной)
    }
    
    shift_types = ['1', '2', 'О', 'В']  # дневная, ночная, отсыпной, выходной
    
    current_date = start_date
    for day in range(days):
        for shift_letter in ['А', 'Б', 'В', 'Г']:
            shift_type = shift_types[(initial_state[shift_letter] + day) % 4]
            
            if shift_type == '1':
                time_start = '08:00'
                time_end = '20:00'
            elif shift_type == '2':
                time_start = '20:00'
                time_end = '08:00'
            else:
                time_start = time_end = None
                
            shifts[shift_letter].append({
                'date': current_date,
                'type': shift_type,
                'time_start': time_start,
                'time_end': time_end
            })
        current_date += timedelta(days=1)
    
    return shifts

@main.route('/shifts')
@login_required
def shifts():
    # Устанавливаем начальную дату как 1 ноября 2024
    start_date = datetime(2024, 11, 1).date()
    days = 30  # Показывать график на 30 дней вперед
    
    schedule = calculate_shift_schedule(start_date, days)
    
    return render_template('shifts.html', schedule=schedule, start_date=start_date, timedelta=timedelta)

@main.route('/update_shift_data', methods=['POST'])
@login_required
def update_shift_data():
    parameter_id = request.json.get('parameter_id')
    if parameter_id:
        session['selected_parameter_id'] = parameter_id
        
        # Получаем даты из сессии
        end_date = session.get('end_date', datetime.now())
        start_date = session.get('start_date', end_date - timedelta(days=7))
        
        # Получаем новые данные для выбранного параметра
        shift_data = db.session.query(
            func.sum(HourlyValues.Value).label('total_value'),
            func.date(HourlyValues.DateTime).label('date'),
            func.extract('hour', HourlyValues.DateTime).label('hour')
        ).filter(
            and_(
                HourlyValues.DateTime.between(start_date, end_date),
                HourlyValues.CodeId == parameter_id
            )
        ).group_by(
            func.date(HourlyValues.DateTime),
            func.extract('hour', HourlyValues.DateTime)
        ).all()
        
        # Агрегируем данные по сменам
        shift_totals = {'А': 0, 'Б': 0, 'В': 0, 'Г': 0}
        
        for value, date, hour in shift_data:
            if isinstance(date, str):
                date = datetime.strptime(date, '%Y-%m-%d').date()
            dt = datetime.combine(date, datetime.min.time().replace(hour=int(hour)))
            shift = get_shift_for_datetime(dt)
            if shift and value:
                shift_totals[shift] += value
        
        return jsonify(shift_totals)
    
    return jsonify({'error': 'Parameter ID not provided'}), 400


@main.route('/export_excel', methods=['POST'])
def export_excel():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Отсутствуют данные запроса'}), 400
            
        start_date = datetime.strptime(data.get('start_date'), '%Y-%m-%dT%H:%M')
        end_date = datetime.strptime(data.get('end_date'), '%Y-%m-%dT%H:%M')
        page_id = data.get('page_id')
        
        if not all([start_date, end_date, page_id]):
            return jsonify({'error': 'Отсутствуют обязательные параметры'}), 400

        # Получаем настройки страницы
        page_settings = PageSettings.query.get_or_404(page_id)
        
        # Получаем параметры в нужном порядке
        parameters = []
        for param_id in page_settings.page_ids:
            param = Parameters.query.get(param_id)
            if param:
                parameters.append(param)
        
        # Получаем данные из БД
        query_result = db.session.query(
            HourlyValues.DateTime,
            HourlyValues.Value,
            Parameters.Name,
            Parameters.CodeId
        ).join(
            Parameters,
            HourlyValues.CodeId == Parameters.CodeId
        ).filter(
            HourlyValues.DateTime.between(start_date, end_date),
            HourlyValues.CodeId.in_(page_settings.page_ids)
        ).order_by(
            HourlyValues.DateTime
        ).all()

        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Данные"
        
        # Формируем заголовки в нужном порядке
        headers = ['Дата и время']
        headers.extend([param.Name for param in parameters])
        ws.append(headers)
        
        # Выделяем заголовки цветом
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
        
        # Группируем данные по дате
        grouped_data = {}
        for record in query_result:
            dt = convert_time(record.DateTime)
            if dt not in grouped_data:
                grouped_data[dt] = {}
            grouped_data[dt][record.CodeId] = record.Value
            
        # Записываем данные
        for dt in sorted(grouped_data.keys()):
            # Преобразуем дату в числовой формат Excel
            excel_date = (dt.timestamp()) / 86400 + 25569 # Добавляем 10800 секунд (3 часа) для московского времени
            row = [excel_date]
            # Добавляем значения в том же порядке, что и заголовки
            for param in parameters:
                row.append(grouped_data[dt].get(param.CodeId, ''))
            ws.append(row)
        
        # Устанавливаем формат даты для первого столбца
        for cell in ws['A'][1:]:
            cell.number_format = 'dd.mm.yyyy hh:mm'
            
        # Настраиваем ширину столбцов
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
        # Закрепляем первую строку и столбец
        ws.freeze_panes = 'B2'
            
        # Сохраняем в буфер
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{page_settings.page_name}_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
        )
        
    except ValueError as e:
        return jsonify({'error': f'Ошибка формата данных: {str(e)}'}), 400
    except Exception as e:
        app.logger.error(f'Ошибка при экспорте в Excel: {str(e)}')
        return jsonify({'error': 'Внутренняя ошибка сервера'}), 500
    

@main.route('/create_page', methods=['GET', 'POST'])
def create_page():
    if request.method == 'POST':
        page_name = request.form['page_name']
        id_list = request.form.get('parameters_order', '')
        parent_id = request.form.get('parent_id')

        if not id_list:
            return jsonify({
                'success': False,
                'error': 'Выберите хотя бы один параметр'
            })

        try:
            # Создаем новую страницу
            new_page = PageSettings(
                page_name=page_name,
                id_list=id_list
            )
            db.session.add(new_page)
            db.session.flush()  # Получаем id новой страницы

            # Создаем иерархическую связь, если указан родитель
            if parent_id and parent_id.isdigit():
                hierarchy = PageHierarchy(
                    parent_id=int(parent_id),
                    child_id=new_page.id,
                    order=len(PageHierarchy.query.filter_by(parent_id=parent_id).all())
                )
                db.session.add(hierarchy)
            
            db.session.commit()
            return jsonify({
                'success': True,
                'redirect_url': url_for('main.manage_pages')
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({
                'success': False,
                'error': str(e)
            })
    
    parameters = Parameters.query.order_by(Parameters.CodeId).all()
    pages = PageSettings.query.all()
    return render_template('create_page.html', parameters=parameters, pages=pages)

@main.route('/manage_pages', methods=['GET', 'POST'])
def manage_pages():
    if request.method == 'POST':
        action = request.form.get('action')
        page_id = request.form.get('page_id')
        
        if action == 'delete':
            try:
                page = PageSettings.query.get_or_404(page_id)
                # Удаляем все связи с этой страницей
                PageHierarchy.query.filter(
                    (PageHierarchy.parent_id == page_id) | 
                    (PageHierarchy.child_id == page_id)
                ).delete()
                db.session.delete(page)
                db.session.commit()
                return jsonify({'success': True})
            except Exception as e:
                db.session.rollback()
                return jsonify({'success': False, 'error': str(e)})
            
        elif action == 'update':
            page = PageSettings.query.get_or_404(page_id)
            page.page_name = request.form.get('page_name')
            new_ids = request.form.get('parameters_order', '')
            new_parent_id = request.form.get('parent_id')
            
            if not new_ids:
                return jsonify({'success': False, 'error': 'Выберите хотя бы один параметр'})
            
            try:
                # Обновляем id_list
                page.id_list = new_ids
                
                # Обновляем иерархию
                current_hierarchy = PageHierarchy.query.filter_by(child_id=page_id).first()
                
                if new_parent_id and new_parent_id.isdigit():
                    new_parent_id = int(new_parent_id)
                    if current_hierarchy:
                        if current_hierarchy.parent_id != new_parent_id:
                            current_hierarchy.parent_id = new_parent_id
                    else:
                        new_hierarchy = PageHierarchy(
                            parent_id=new_parent_id,
                            child_id=page_id,
                            order=len(PageHierarchy.query.filter_by(parent_id=new_parent_id).all())
                        )
                        db.session.add(new_hierarchy)
                else:
                    if current_hierarchy:
                        db.session.delete(current_hierarchy)
                
                # Обновляем типы агрегации
                for param_id in new_ids.split(','):
                    param = Parameters.query.get(param_id)
                    if param:
                        new_agg_type = request.form.get(f'aggregation_type_{param_id}')
                        if new_agg_type in ['avg', 'sum', 'min', 'max', 'weighted_avg']:
                            param.aggregation_type = new_agg_type
                            if new_agg_type == 'weighted_avg':
                                weight_param_id = request.form.get(f'weight_parameter_{param_id}')
                                if weight_param_id and weight_param_id.isdigit():
                                    param.weight_parameter_id = int(weight_param_id)
                            else:
                                param.weight_parameter_id = None
                
                db.session.commit()
                return jsonify({'success': True})
            except Exception as e:
                db.session.rollback()
                return jsonify({'success': False, 'error': str(e)})
    
    # Получаем иерархию страниц
    root_pages = get_page_hierarchy()
    parameters = Parameters.query.order_by(Parameters.CodeId).all()
    all_pages = PageSettings.query.all()
    return render_template('manage_pages.html', 
                         root_pages=root_pages, 
                         parameters=parameters,
                         all_pages=all_pages)

@main.route('/manage_aggregation', methods=['GET', 'POST'])
def manage_aggregation():
    parameters = Parameters.query.order_by(Parameters.CodeId).all()
    
    if request.method == 'POST':
        param_id = request.form.get('param_id')
        new_agg_type = request.form.get('aggregation_type')
        weight_param_id = request.form.get('weight_parameter_id')
        
        param = Parameters.query.get_or_404(param_id)
        if new_agg_type in ['avg', 'sum', 'min', 'max', 'weighted_avg']:
            param.aggregation_type = new_agg_type
            if new_agg_type == 'weighted_avg' and weight_param_id:
                param.weight_parameter_id = int(weight_param_id)
            else:
                param.weight_parameter_id = None
            db.session.commit()
            return jsonify({'success': True})
        
        return jsonify({'success': False, 'error': 'Неверный тип агрегации'})
        
    return render_template('manage_aggregation.html', parameters=parameters)

def get_page_hierarchy():
    # Получаем все страницы первого уровня (без родителя)
    root_pages = db.session.query(PageSettings)\
        .outerjoin(PageHierarchy, PageHierarchy.child_id == PageSettings.id)\
        .filter(PageHierarchy.parent_id == None)\
        .order_by(PageHierarchy.order)\
        .all()
    
    # Рекурсивно получаем дочерние страницы
    def get_children(page):
        children = db.session.query(PageSettings)\
            .join(PageHierarchy, PageHierarchy.child_id == PageSettings.id)\
            .filter(PageHierarchy.parent_id == page.id)\
            .order_by(PageHierarchy.order)\
            .all()
        
        for child in children:
            child.children = get_children(child)
        return children
    
    # Добавляем дочерние страницы к корневым
    for page in root_pages:
        page.children = get_children(page)
    
    return root_pages

@main.context_processor
def inject_pages():
    return dict(root_pages=get_page_hierarchy())

@main.route('/update_page/<int:page_id>', methods=['POST'])
def update_page(page_id):
    page = PageSettings.query.get_or_404(page_id)
    new_parent_id = request.form.get('parent_id')
    
    try:
        # Получаем текущего родителя
        current_hierarchy = PageHierarchy.query.filter_by(child_id=page_id).first()
        old_parent_id = current_hierarchy.parent_id if current_hierarchy else None

        # Если родитель изменился
        if str(old_parent_id) != str(new_parent_id):
            # Очищаем кэш старого родителя
            if old_parent_id:
                old_parent = PageSettings.query.get(old_parent_id)
                if old_parent:
                    old_parent.clear_children_cache()

            # Очищаем кэш нового родителя
            if new_parent_id and new_parent_id.isdigit():
                new_parent = PageSettings.query.get(int(new_parent_id))
                if new_parent:
                    new_parent.clear_children_cache()

            # Обновляем иерархию
            if current_hierarchy:
                if new_parent_id and new_parent_id.isdigit():
                    current_hierarchy.parent_id = int(new_parent_id)
                else:
                    db.session.delete(current_hierarchy)
            elif new_parent_id and new_parent_id.isdigit():
                hierarchy = PageHierarchy(
                    parent_id=int(new_parent_id),
                    child_id=page_id,
                    order=len(PageHierarchy.query.filter_by(parent_id=new_parent_id).all())
                )
                db.session.add(hierarchy)

        # Обновляем основные данные страницы
        page.page_name = request.form.get('page_name')
        page.id_list = request.form.get('parameters_order', '')
        
        # Проверяем наличие параметров
        if not page.id_list:
            return jsonify({
                'success': False,
                'error': 'Выберите хотя бы один параметр'
            })
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        })

@main.route('/delete_page/<int:page_id>', methods=['POST'])
def delete_page(page_id):
    try:
        page = PageSettings.query.get_or_404(page_id)
        
        # Получаем и очищаем кэш родительской страницы
        parent = page.parent
        if parent:
            parent.clear_children_cache()
            
        # Получаем все дочерние страницы рекурсивно
        def get_all_children(page):
            children = page.children
            all_children = list(children)
            for child in children:
                all_children.extend(get_all_children(child))
            return all_children
            
        # Очищаем кэш всех дочерних страниц
        children = get_all_children(page)
        for child in children:
            child.clear_children_cache()
            
        # Удаляем все связи
        PageHierarchy.query.filter(
            (PageHierarchy.parent_id == page_id) |
            (PageHierarchy.child_id == page_id)
        ).delete()
        
        # Удаляем саму страницу
        db.session.delete(page)
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        })

@main.route('/reorder_pages', methods=['POST'])
def reorder_pages():
    try:
        data = request.get_json()
        page_id = data.get('page_id')
        parent_id = data.get('parent_id')
        page_order = data.get('page_order', [])
        
        # Получаем перемещаемую страницу
        page = PageSettings.query.get_or_404(page_id)
        
        # Получаем текущую иерархию
        current_hierarchy = PageHierarchy.query.filter_by(child_id=page_id).first()
        
        # Если есть старый родитель, очищаем его кэш
        if current_hierarchy and current_hierarchy.parent_id:
            old_parent = PageSettings.query.get(current_hierarchy.parent_id)
            if old_parent:
                old_parent.clear_children_cache()
        
        # Обновляем или создаем новую иерархию
        if parent_id:
            if current_hierarchy:
                current_hierarchy.parent_id = parent_id
            else:
                new_hierarchy = PageHierarchy(
                    parent_id=parent_id,
                    child_id=page_id,
                    order=len(page_order) - 1
                )
                db.session.add(new_hierarchy)
            
            # Очищаем кэш нового родителя
            new_parent = PageSettings.query.get(parent_id)
            if new_parent:
                new_parent.clear_children_cache()
        else:
            # Если страница перемещается в корень
            if current_hierarchy:
                db.session.delete(current_hierarchy)
        
        # Обновляем порядок страниц
        for index, p_id in enumerate(page_order):
            h = PageHierarchy.query.filter_by(
                parent_id=parent_id,
                child_id=p_id
            ).first()
            if h:
                h.order = index
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        })